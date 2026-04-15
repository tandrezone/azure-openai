"""
Azure OpenAI – Excel Product Description Generator

This example reads product data from an Excel file and uses Azure OpenAI to
generate a marketing description for each product based on its attributes.

Prerequisites:
    - An Azure OpenAI resource with a deployed model
    - The openai and openpyxl Python packages
      (pip install -r examples/requirements.txt)
    - Environment variables set for authentication (see below)

Environment Variables:
    AZURE_OPENAI_ENDPOINT:        Your Azure OpenAI resource endpoint
                                  (e.g. https://<your-resource>.openai.azure.com/)
    AZURE_OPENAI_API_KEY:         Your Azure OpenAI API key
    AZURE_OPENAI_DEPLOYMENT_NAME: The name of your deployed model (e.g. gpt-4o)
    AZURE_OPENAI_API_VERSION:     API version (default: 2024-12-01-preview)

Usage:
    python examples/excel_product_description.py examples/sample_products.xlsx

    Optionally write results to a new Excel file:
    python examples/excel_product_description.py examples/sample_products.xlsx --output descriptions.xlsx
"""

from __future__ import annotations

import argparse
import os
import sys

import openpyxl
from openai import AzureOpenAI


def get_env_var(name: str, default: str | None = None) -> str:
    """Retrieve a required environment variable or exit with an error."""
    value = os.environ.get(name, default)
    if value is None:
        print(f"Error: environment variable '{name}' is not set.")
        sys.exit(1)
    return value


def read_products(filepath: str) -> list[dict[str, str]]:
    """Read products from the first sheet of an Excel workbook.

    The first row is treated as headers. Each subsequent row becomes a
    dictionary mapping header names to cell values.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        print("Error: the Excel file must have a header row and at least one data row.")
        sys.exit(1)

    headers = [str(h).strip() for h in rows[0] if h is not None]
    products: list[dict[str, str]] = []
    for row in rows[1:]:
        product = {headers[i]: str(v) if v is not None else "" for i, v in enumerate(row) if i < len(headers)}
        if any(product.values()):
            products.append(product)

    return products


def build_prompt(product: dict[str, str]) -> str:
    """Build a user prompt that lists product attributes."""
    attributes = "\n".join(f"- {key}: {value}" for key, value in product.items() if value)
    return (
        "Based on the following product attributes, write a short, compelling "
        "marketing description (2-3 sentences) for this product. "
        "Highlight key selling points and appeal to the target audience.\n\n"
        f"Product attributes:\n{attributes}"
    )


def generate_description(client: AzureOpenAI, deployment: str, product: dict[str, str]) -> str:
    """Send a product's attributes to Azure OpenAI and return the description."""
    messages = [
        {
            "role": "system",
            "content": (
                "You are a professional copywriter specializing in product "
                "descriptions. Write concise and engaging marketing copy."
            ),
        },
        {
            "role": "user",
            "content": build_prompt(product),
        },
    ]

    response = client.chat.completions.create(
        model=deployment,
        messages=messages,
        temperature=0.7,
        max_tokens=256,
    )
    return response.choices[0].message.content.strip()


def save_results(filepath: str, products: list[dict[str, str]], descriptions: list[str]) -> None:
    """Write the products and their generated descriptions to a new Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Descriptions"

    headers = list(products[0].keys()) + ["Generated Description"]
    ws.append(headers)

    for product, description in zip(products, descriptions):
        row = list(product.values()) + [description]
        ws.append(row)

    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

    wb.save(filepath)


def parse_args() -> argparse.Namespace:
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Generate product descriptions from an Excel file using Azure OpenAI.",
    )
    parser.add_argument(
        "excel_file",
        help="Path to the Excel (.xlsx) file containing product data.",
    )
    parser.add_argument(
        "--output", "-o",
        default=None,
        help="Optional output Excel file path to save results. "
             "If not provided, descriptions are printed to the console only.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    # ── Azure OpenAI client setup ──────────────────────────────────────
    endpoint = get_env_var("AZURE_OPENAI_ENDPOINT")
    api_key = get_env_var("AZURE_OPENAI_API_KEY")
    deployment = get_env_var("AZURE_OPENAI_DEPLOYMENT_NAME")
    api_version = get_env_var("AZURE_OPENAI_API_VERSION", "2024-12-01-preview")

    client = AzureOpenAI(
        azure_endpoint=endpoint,
        api_key=api_key,
        api_version=api_version,
    )

    # ── Read Excel data ────────────────────────────────────────────────
    print(f"Reading products from '{args.excel_file}'...\n")
    products = read_products(args.excel_file)
    print(f"Found {len(products)} product(s).\n")

    # ── Generate descriptions ──────────────────────────────────────────
    descriptions: list[str] = []
    for i, product in enumerate(products, start=1):
        name = product.get("Product Name", product.get(list(product.keys())[0], f"Product {i}"))
        print(f"[{i}/{len(products)}] Generating description for '{name}'...")
        description = generate_description(client, deployment, product)
        descriptions.append(description)
        print(f"  → {description}\n")

    # ── Save results ───────────────────────────────────────────────────
    if args.output:
        save_results(args.output, products, descriptions)
        print(f"Results saved to '{args.output}'.")


if __name__ == "__main__":
    main()
